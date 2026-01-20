"""
Excel export functionality for HUD Financing Platform
Generates comprehensive workbooks with all deal data
"""
from io import BytesIO
from typing import Dict, List, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import pandas as pd

from .deal import Deal
from .cashflows import generate_cashflows, CashflowResult


# Ascendra colors for Excel
COLORS = {
    "dark_bg": "0A1929",
    "cyan": "4CC9F0",
    "mint": "06FFA5",
    "orange": "FFA15A",
    "red": "EF553B",
    "white": "FFFFFF",
    "light_gray": "B0BEC5",
}


def create_excel_workbook(
    deal: Deal,
    sofr_curve: List[float],
    exit_month: int,
    include_scenarios: bool = True,
    include_sensitivity: bool = True,
) -> Optional[BytesIO]:
    """
    Create comprehensive Excel workbook for deal

    Args:
        deal: Deal object
        sofr_curve: SOFR curve
        exit_month: Exit month
        include_scenarios: Include scenario analysis
        include_sensitivity: Include sensitivity tables

    Returns:
        BytesIO buffer with Excel file, or None if openpyxl not available
    """
    if not OPENPYXL_AVAILABLE:
        return None

    wb = Workbook()

    # Generate cashflows
    results = generate_cashflows(deal, sofr_curve, exit_month)

    # Create sheets
    _create_summary_sheet(wb, deal, results, sofr_curve[0])
    _create_cashflow_sheet(wb, results, "Sponsor Cashflows", "sponsor")
    _create_cashflow_sheet(wb, results, "A-Piece Cashflows", "A")
    _create_cashflow_sheet(wb, results, "B-Piece Cashflows", "B")
    _create_cashflow_sheet(wb, results, "C-Piece Cashflows", "C")

    if include_scenarios:
        _create_scenario_sheet(wb, deal, sofr_curve)

    _create_assumptions_sheet(wb, deal, sofr_curve[0])

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def _apply_header_style(cell):
    """Apply header styling to cell"""
    cell.font = Font(bold=True, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["cyan"], end_color=COLORS["cyan"], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_number_format(cell, format_type: str = "number"):
    """Apply number formatting"""
    formats = {
        "number": "#,##0",
        "currency": "$#,##0",
        "percent": "0.00%",
        "decimal": "0.00",
    }
    cell.number_format = formats.get(format_type, "#,##0")


def _create_summary_sheet(
    wb: Workbook,
    deal: Deal,
    results: Dict[str, CashflowResult],
    current_sofr: float,
):
    """Create summary sheet"""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws["A1"] = "HUD FINANCING DEAL SUMMARY"
    ws["A1"].font = Font(bold=True, size=16, color=COLORS["cyan"])
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color=COLORS["light_gray"])

    # Deal Overview Section
    row = 4
    ws[f"A{row}"] = "DEAL OVERVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    overview_data = [
        ("Deal Name", deal.deal_name),
        ("Property Value", deal.property_value),
        ("Loan Amount", deal.loan_amount),
        ("LTV", deal.ltv),
        ("Term (months)", deal.term_months),
        ("Expected HUD Month", deal.expected_hud_month),
    ]

    row += 1
    for label, value in overview_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        if isinstance(value, float) and value < 1:
            _apply_number_format(ws[f"B{row}"], "percent")
        elif isinstance(value, (int, float)) and value > 1000:
            _apply_number_format(ws[f"B{row}"], "currency")
        row += 1

    # Key Metrics Section
    row += 1
    ws[f"A{row}"] = "KEY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    sponsor = results.get("sponsor")
    if sponsor:
        metrics_data = [
            ("Sponsor IRR", sponsor.irr),
            ("Sponsor MOIC", sponsor.moic),
            ("Total Profit", sponsor.total_profit),
            ("Blended Cost of Capital", deal.get_blended_cost_of_capital(current_sofr)),
            ("Spread Capture", deal.get_borrower_rate(current_sofr) - deal.get_blended_cost_of_capital(current_sofr)),
        ]

        row += 1
        for label, value in metrics_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            if "IRR" in label or "Spread" in label or "Cost" in label:
                _apply_number_format(ws[f"B{row}"], "percent")
            elif "MOIC" in label:
                _apply_number_format(ws[f"B{row}"], "decimal")
            else:
                _apply_number_format(ws[f"B{row}"], "currency")
            row += 1

    # Tranche Returns Section
    row += 1
    ws[f"A{row}"] = "TRANCHE RETURNS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    headers = ["Tranche", "Amount", "% of Loan", "IRR", "MOIC", "Profit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        _apply_header_style(cell)

    for tranche in deal.tranches:
        row += 1
        t_result = results.get(tranche.tranche_type.value)
        amount = deal.get_tranche_amount(tranche)

        ws.cell(row=row, column=1, value=tranche.name)
        ws.cell(row=row, column=2, value=amount)
        _apply_number_format(ws.cell(row=row, column=2), "currency")

        ws.cell(row=row, column=3, value=tranche.percentage)
        _apply_number_format(ws.cell(row=row, column=3), "percent")

        if t_result:
            ws.cell(row=row, column=4, value=t_result.irr)
            _apply_number_format(ws.cell(row=row, column=4), "percent")

            ws.cell(row=row, column=5, value=t_result.moic)
            _apply_number_format(ws.cell(row=row, column=5), "decimal")

            ws.cell(row=row, column=6, value=t_result.total_profit)
            _apply_number_format(ws.cell(row=row, column=6), "currency")

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 15


def _create_cashflow_sheet(
    wb: Workbook,
    results: Dict[str, CashflowResult],
    sheet_name: str,
    result_key: str,
):
    """Create cashflow sheet for specific stakeholder"""
    cf = results.get(result_key)
    if not cf:
        return

    ws = wb.create_sheet(title=sheet_name)

    # Headers
    headers = ["Month", "Principal", "Interest", "Fees", "Net Cashflow", "Cumulative"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    # Data
    cumulative = 0
    for i, month in enumerate(cf.months):
        row = i + 2
        cumulative += cf.total_flows[i]

        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=cf.principal_flows[i])
        ws.cell(row=row, column=3, value=cf.interest_flows[i])
        ws.cell(row=row, column=4, value=cf.fee_flows[i])
        ws.cell(row=row, column=5, value=cf.total_flows[i])
        ws.cell(row=row, column=6, value=cumulative)

        for col in range(2, 7):
            _apply_number_format(ws.cell(row=row, column=col), "currency")

    # Summary row
    row = len(cf.months) + 3
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)

    ws.cell(row=row, column=2, value=sum(cf.principal_flows))
    ws.cell(row=row, column=3, value=sum(cf.interest_flows))
    ws.cell(row=row, column=4, value=sum(cf.fee_flows))
    ws.cell(row=row, column=5, value=sum(cf.total_flows))

    for col in range(2, 6):
        _apply_number_format(ws.cell(row=row, column=col), "currency")
        ws.cell(row=row, column=col).font = Font(bold=True)

    # Metrics
    row += 2
    ws.cell(row=row, column=1, value="IRR")
    ws.cell(row=row, column=2, value=cf.irr)
    _apply_number_format(ws.cell(row=row, column=2), "percent")

    ws.cell(row=row + 1, column=1, value="MOIC")
    ws.cell(row=row + 1, column=2, value=cf.moic)
    _apply_number_format(ws.cell(row=row + 1, column=2), "decimal")

    # Column widths
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 15


def _create_scenario_sheet(
    wb: Workbook,
    deal: Deal,
    sofr_curve: List[float],
):
    """Create scenario analysis sheet"""
    from .scenarios import get_standard_scenarios, run_scenarios

    ws = wb.create_sheet(title="Scenarios")

    scenarios = get_standard_scenarios(deal)
    results = run_scenarios(deal, scenarios, sofr_curve)

    # Headers
    headers = ["Scenario", "Exit Month", "Extension", "Sponsor IRR", "Sponsor MOIC", "Profit", "A IRR", "B IRR"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    # Data
    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r.scenario.name)
        ws.cell(row=i, column=2, value=r.scenario.exit_month)
        ws.cell(row=i, column=3, value="Yes" if r.scenario.has_extension else "No")

        ws.cell(row=i, column=4, value=r.sponsor_irr)
        _apply_number_format(ws.cell(row=i, column=4), "percent")

        ws.cell(row=i, column=5, value=r.sponsor_moic)
        _apply_number_format(ws.cell(row=i, column=5), "decimal")

        ws.cell(row=i, column=6, value=r.sponsor_profit)
        _apply_number_format(ws.cell(row=i, column=6), "currency")

        ws.cell(row=i, column=7, value=r.a_irr)
        _apply_number_format(ws.cell(row=i, column=7), "percent")

        ws.cell(row=i, column=8, value=r.b_irr)
        _apply_number_format(ws.cell(row=i, column=8), "percent")

    # Column widths
    ws.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 12


def _create_assumptions_sheet(
    wb: Workbook,
    deal: Deal,
    current_sofr: float,
):
    """Create assumptions sheet"""
    ws = wb.create_sheet(title="Assumptions")

    ws["A1"] = "DEAL ASSUMPTIONS"
    ws["A1"].font = Font(bold=True, size=14)

    assumptions = [
        ("", ""),
        ("Property & Loan", ""),
        ("Property Value", deal.property_value),
        ("Loan Amount", deal.loan_amount),
        ("LTV", deal.ltv),
        ("Term (months)", deal.term_months),
        ("Expected HUD Month", deal.expected_hud_month),
        ("", ""),
        ("Interest Rates", ""),
        ("Current SOFR", current_sofr),
        ("Borrower Spread", deal.borrower_spread),
        ("Borrower Rate", deal.get_borrower_rate(current_sofr)),
        ("", ""),
        ("Capital Stack", ""),
        ("A-Piece %", deal.tranches[0].percentage if len(deal.tranches) > 0 else 0),
        ("A-Piece Spread", deal.tranches[0].spread if len(deal.tranches) > 0 else 0),
        ("B-Piece %", deal.tranches[1].percentage if len(deal.tranches) > 1 else 0),
        ("B-Piece Spread", deal.tranches[1].spread if len(deal.tranches) > 1 else 0),
        ("C-Piece %", deal.tranches[2].percentage if len(deal.tranches) > 2 else 0),
        ("C-Piece Rate", deal.tranches[2].spread if len(deal.tranches) > 2 else 0),
        ("", ""),
        ("Fees", ""),
        ("Origination Fee", deal.fees.origination_fee),
        ("Exit Fee", deal.fees.exit_fee),
        ("Extension Fee", deal.fees.extension_fee),
    ]

    for row, (label, value) in enumerate(assumptions, 2):
        ws.cell(row=row, column=1, value=label)
        if label and not label.endswith("&") and value != "":
            ws.cell(row=row, column=2, value=value)
            if isinstance(value, float) and value < 1:
                _apply_number_format(ws.cell(row=row, column=2), "percent")
            elif isinstance(value, (int, float)) and value > 100:
                _apply_number_format(ws.cell(row=row, column=2), "currency")

        # Bold section headers
        if value == "":
            ws.cell(row=row, column=1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def export_cashflows_to_csv(
    results: Dict[str, CashflowResult],
    stakeholder: str = "sponsor",
) -> str:
    """
    Export cashflows to CSV string

    Args:
        results: Cashflow results dict
        stakeholder: Which stakeholder's cashflows to export

    Returns:
        CSV string
    """
    cf = results.get(stakeholder)
    if not cf:
        return ""

    df = pd.DataFrame({
        "Month": cf.months,
        "Principal": cf.principal_flows,
        "Interest": cf.interest_flows,
        "Fees": cf.fee_flows,
        "Net_Cashflow": cf.total_flows,
    })

    return df.to_csv(index=False)


def export_all_cashflows_to_csv(results: Dict[str, CashflowResult]) -> str:
    """Export all stakeholder cashflows to single CSV"""
    sponsor = results.get("sponsor")
    if not sponsor:
        return ""

    df = pd.DataFrame({"Month": sponsor.months})

    for name, key in [("Sponsor", "sponsor"), ("A", "A"), ("B", "B"), ("C", "C")]:
        cf = results.get(key)
        if cf:
            df[f"{name}_Principal"] = cf.principal_flows
            df[f"{name}_Interest"] = cf.interest_flows
            df[f"{name}_Fees"] = cf.fee_flows
            df[f"{name}_Net"] = cf.total_flows

    return df.to_csv(index=False)
